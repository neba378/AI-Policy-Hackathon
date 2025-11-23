"""
Dashboard and export services for Policy Sentinel.
"""
from typing import List, Dict, Any
from collections import defaultdict
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.schemas import AuditItem, Rule


def calculate_dashboard_stats(audit_results: List[AuditItem], rules: List[Rule]) -> Dict[str, Any]:
    """
    Calculate compliance dashboard statistics.
    
    Returns:
        Dashboard data with model rankings, category breakdowns, and overall stats
    """
    # Group results by model
    model_stats = defaultdict(lambda: {"PASS": 0, "FAIL": 0, "total": 0, "avg_confidence": 0.0})
    category_stats = defaultdict(lambda: {"PASS": 0, "FAIL": 0})
    
    for item in audit_results:
        status = item.evidence.status
        confidence = item.evidence.confidence
        model_stats[item.model_name][status] += 1
        model_stats[item.model_name]["total"] += 1
        model_stats[item.model_name]["avg_confidence"] += confidence
        category_stats[item.rule_category][status] += 1
    
    # Calculate compliance scores and average confidence
    model_rankings = []
    for model_name, stats in model_stats.items():
        total = stats["total"]
        compliance_score = (stats["PASS"] / total) * 100 if total > 0 else 0.0
        avg_confidence = stats["avg_confidence"] / total if total > 0 else 0.0
        
        model_rankings.append({
            "model_name": model_name,
            "compliance_score": round(compliance_score, 1),
            "avg_confidence": round(avg_confidence, 1),
            "pass_count": stats["PASS"],
            "fail_count": stats["FAIL"],
            "total_rules": stats["total"]
        })
    
    # Sort by compliance score
    model_rankings.sort(key=lambda x: x["compliance_score"], reverse=True)
    
    # Calculate category breakdown
    category_breakdown = []
    for category, stats in category_stats.items():
        total = stats["PASS"] + stats["FAIL"]
        pass_rate = (stats["PASS"] / total) * 100 if total > 0 else 0.0
        
        category_breakdown.append({
            "category": category,
            "pass_rate": round(pass_rate, 1),
            "pass_count": stats["PASS"],
            "fail_count": stats["FAIL"]
        })
    
    # Overall statistics
    total_checks = len(audit_results)
    total_pass = sum(1 for item in audit_results if item.evidence.status == "PASS")
    total_fail = sum(1 for item in audit_results if item.evidence.status == "FAIL")
    overall_compliance = (total_pass / total_checks) * 100 if total_checks > 0 else 0.0
    avg_overall_confidence = sum(item.evidence.confidence for item in audit_results) / total_checks if total_checks > 0 else 0.0
    
    return {
        "overall_compliance": round(overall_compliance, 1),
        "avg_confidence": round(avg_overall_confidence, 1),
        "total_checks": total_checks,
        "total_pass": total_pass,
        "total_fail": total_fail,
        "model_rankings": model_rankings,
        "category_breakdown": category_breakdown,
        "best_model": model_rankings[0]["model_name"] if model_rankings else None,
        "worst_model": model_rankings[-1]["model_name"] if model_rankings else None
    }


def generate_excel_report(
    policy_name: str,
    rules: List[Rule],
    audit_results: List[AuditItem],
    dashboard_stats: Dict[str, Any]
) -> BytesIO:
    """
    Generate an Excel report with multiple sheets.
    
    Returns:
        BytesIO object containing the Excel file
    """
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Sheet 1: Dashboard Summary
    ws_dashboard = wb.create_sheet("Dashboard")
    
    # Title
    ws_dashboard['A1'] = "Policy Sentinel - Compliance Report"
    ws_dashboard['A1'].font = Font(size=16, bold=True)
    ws_dashboard['A2'] = f"Policy: {policy_name}"
    ws_dashboard['A2'].font = Font(size=12, italic=True)
    
    # Overall Stats
    ws_dashboard['A4'] = "Overall Compliance"
    ws_dashboard['A4'].font = Font(bold=True)
    ws_dashboard['B4'] = f"{dashboard_stats['overall_compliance']}%"
    ws_dashboard['B4'].font = Font(size=14, bold=True, color="00B050")
    
    ws_dashboard['A5'] = "Total Checks"
    ws_dashboard['B5'] = dashboard_stats['total_checks']
    ws_dashboard['A6'] = "Passed"
    ws_dashboard['B6'] = dashboard_stats['total_pass']
    ws_dashboard['A8'] = "Failed"
    ws_dashboard['B8'] = dashboard_stats['total_fail']
    ws_dashboard['A9'] = "Avg Confidence"
    ws_dashboard['B9'] = f"{dashboard_stats['avg_confidence']}%"
    
    # Model Rankings
    ws_dashboard['A11'] = "Model Rankings"
    ws_dashboard['A11'].font = Font(bold=True, size=12)
    
    headers = ['Rank', 'Model', 'Compliance Score', 'Avg Confidence', 'Pass', 'Fail']
    for col, header in enumerate(headers, 1):
        cell = ws_dashboard.cell(row=12, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    for idx, model in enumerate(dashboard_stats['model_rankings'], 1):
        ws_dashboard.cell(row=12+idx, column=1, value=idx)
        ws_dashboard.cell(row=12+idx, column=2, value=model['model_name'])
        ws_dashboard.cell(row=12+idx, column=3, value=f"{model['compliance_score']}%")
        ws_dashboard.cell(row=12+idx, column=4, value=f"{model['avg_confidence']}%")
        ws_dashboard.cell(row=12+idx, column=5, value=model['pass_count'])
        ws_dashboard.cell(row=12+idx, column=6, value=model['fail_count'])
    
    # Adjust column widths
    ws_dashboard.column_dimensions['A'].width = 20
    ws_dashboard.column_dimensions['B'].width = 30
    ws_dashboard.column_dimensions['C'].width = 18
    ws_dashboard.column_dimensions['D'].width = 10
    ws_dashboard.column_dimensions['E'].width = 10
    ws_dashboard.column_dimensions['F'].width = 10
    
    # Sheet 2: Detailed Results
    ws_details = wb.create_sheet("Detailed Results")
    
    detail_headers = ['Model', 'Rule ID', 'Category', 'Question', 'Status', 'Confidence', 'Evidence Quote', 'Reason']
    for col, header in enumerate(detail_headers, 1):
        cell = ws_details.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    for idx, item in enumerate(audit_results, 2):
        ws_details.cell(row=idx, column=1, value=item.model_name)
        ws_details.cell(row=idx, column=2, value=item.rule_id)
        ws_details.cell(row=idx, column=3, value=item.rule_category)
        ws_details.cell(row=idx, column=4, value=item.rule_question)
        
        status_cell = ws_details.cell(row=idx, column=5, value=item.evidence.status)
        # Color code status
        if item.evidence.status == "PASS":
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        ws_details.cell(row=idx, column=6, value=f"{item.evidence.confidence}%")
        ws_details.cell(row=idx, column=7, value=item.evidence.quote)
        ws_details.cell(row=idx, column=8, value=item.evidence.reason)
    
    # Adjust column widths
    ws_details.column_dimensions['A'].width = 20
    ws_details.column_dimensions['B'].width = 10
    ws_details.column_dimensions['C'].width = 15
    ws_details.column_dimensions['D'].width = 50
    ws_details.column_dimensions['E'].width = 12
    ws_details.column_dimensions['F'].width = 12
    ws_details.column_dimensions['G'].width = 60
    ws_details.column_dimensions['H'].width = 40
    
    # Sheet 3: Rules
    ws_rules = wb.create_sheet("Policy Rules")
    
    rule_headers = ['ID', 'Category', 'Question']
    for col, header in enumerate(rule_headers, 1):
        cell = ws_rules.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    for idx, rule in enumerate(rules, 2):
        ws_rules.cell(row=idx, column=1, value=rule.id)
        ws_rules.cell(row=idx, column=2, value=rule.category)
        ws_rules.cell(row=idx, column=3, value=rule.question)
    
    ws_rules.column_dimensions['A'].width = 10
    ws_rules.column_dimensions['B'].width = 20
    ws_rules.column_dimensions['C'].width = 70
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file
